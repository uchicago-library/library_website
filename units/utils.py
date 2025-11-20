from base.utils import get_xml_from_directory_api
from django.conf import settings
from django.core.cache import cache
from django.db.models.base import ObjectDoesNotExist
from django.utils.text import slugify
from file_parsing import is_int
from library_website.settings import DEFAULT_UNIT, PUBLIC_HOMEPAGE, PUBLIC_SITE
from openpyxl import Workbook
from site_settings.models import QuickNumberGroup
from wagtail.models import Page, Site
from xml.etree import ElementTree

import csv
import io
import re


def get_quick_nums_dict():
    """
    Get quick numbers dictionary from Wagtail site settings.
    Results are cached for 24 hours.

    Returns:
        dict: Quick numbers in format {slug: [{'label': str, 'number': str, 'link': int|None}]}
    """
    cached = cache.get('quick_nums_dict')
    if cached:
        return cached

    result = {}
    groups = QuickNumberGroup.objects.prefetch_related('numbers').all()

    for group in groups:
        result[group.slug] = [
            {
                'label': num.label,
                'number': num.number,
                'link': num.link.id if num.link else None,
            }
            for num in group.numbers.all()
        ]

    # Cache for 24 hours (86400 seconds)
    cache.set('quick_nums_dict', result, 86400)
    return result


def get_default_unit():
    """
    Get a fallback unit by ID from the config
    if no unit is set.

    Returns:
        UnitPage object
    """
    from units.models import UnitPage
    return UnitPage.objects.live().filter(id=DEFAULT_UNIT)[0]


def get_quick_num_or_link(dic):
    """
    Get the quick numbers or a link.

    Args:
        dic: A dictionary with the keys "label", "number"
        and "link", where values for label and number are
        strings. Values for link are either None or an
        integer representing a page ID.

    Returns:
        A tuple where the first item in either 0 or 1. 0
        represents a phone number and 1 represents a link.
        The second item in the tuple is a string label
        for display. If the first item is 0 then the third
        item is a string representing a phone number, if
        the first item is 1, then the thrid  is a string
        representing a url, e.g.:

        (0, 'Main Telephone', '773-702-8740') or
        (1, 'SCRC Contact Form', '/scrc/visiting/contact/')
    """

    assert all(k in dic for k in ['label', 'number', 'link']), \
        'Quick number dictionaries should have the following keys: label, number, link'

    assert dic['label'], 'There must always be a value for label'

    if dic['link']:
        assert is_int(dic['link']), \
            'The link key should contain an integer representing page ID'
        # relative_url gives an absolute url in unit tests, so, use get_url_parts instead
        retval = (1, dic['label'], Page.objects.all().get(id=dic['link']).get_url_parts()[-1])
    elif dic['number']:
        retval = (0, dic['label'], dic['number'])
    else:
        raise ValueError('Missing data: either a number or a link is required')

    return retval


def get_quick_num_html(quick_num):
    """
    Get the html for displaying a single quick number
    or link on the top of the public site directory.

    Args:
        dlist: a list of dictionaries formatted for the
        get_quick_num_or_link function.

        quick_num: a tuple representation of a quick
        number or link to be displayed at the top of
        the directory. The tuple should be of the same
        type returned from get_quick_num_or_link.

    Returns:
        html, representation of a quick number or link.
    """
    if quick_num[0] == 0:
        html = '<td><strong>' + quick_num[1] + '</strong> ' + quick_num[2] +'</td>'
    else:
        html = '<td><strong><a href="' + quick_num[2] + '">' + quick_num[1] + '</a></strong></td>'
    return html


def get_all_quick_nums_html(dlist):
    """
    Get all quick numbers and links for a given department
    to be displayed at the top of the public site unit
    directory listing.

    Args:
        dlist: a list of dictionaries formatted as required
        by get_quick_num_or_link.

    Returns:
        String, the full html to be displayed at the top of
        the department directory on the public site.
    """
    html = ''
    for item in dlist:
        html += get_quick_num_html(get_quick_num_or_link(item))
    return html


def get_quick_nums_for_library_or_dept(request):
    """
    Gets the quick number html for a request by library or 
    department (unit). The following precedence is observed:
    
    1. If "library" is present in the request and config, it is used.
    2. If "department" is present in the request and the config, it 
       is used.
    3. If "department" is in the request but not in the config, 
       unit > location is used.
    4. If nothing is found, the default fallback of library will
       be used.

    Args:
        dlist: a list of dictionaries formatted as required
        by get_quick_num_or_link.

        request: object

    Returns:
        String, the full html to be displayed at the top of
        the department directory chosen by the available data.
    """
    library = request.GET.get('library', None)
    department = request.GET.get('department', None)

    quick_nums = get_quick_nums_dict()

    # Get default fallback
    default_group = QuickNumberGroup.objects.filter(is_default=True).first()
    default_numbers = []
    if default_group and default_group.slug in quick_nums:
        default_numbers = quick_nums[default_group.slug]

    # Start with default
    html = get_all_quick_nums_html(default_numbers)

    # Override with specific library if provided
    if library:
        library_slug = slugify(library)
        if library_slug in quick_nums:
            html = get_all_quick_nums_html(quick_nums[library_slug])
    elif department:
        dept_slug = slugify(department)
        if dept_slug in quick_nums:
            html = get_all_quick_nums_html(quick_nums[dept_slug])
        else:
            # Try looking up by unit location
            try:
                from units.models import UnitIndexPage, UnitPage
                url_path = UnitIndexPage.objects.first().url_path + '/'.join(map(slugify, department.split(' - '))) + '/'
                unitpage = UnitPage.objects.live().get(url_path=url_path)
                location_slug = slugify(unitpage.location)
                if location_slug in quick_nums:
                    html = get_all_quick_nums_html(quick_nums[location_slug])
            except(UnitPage.DoesNotExist, AttributeError):
                pass
    return html


class WagtailUnitsReport:
    """
    Reporting on Wagtail units for the HR department. This class includes
    methods to report on the units that are present in the campus directory but
    not in Wagtail, and it includes general reporting for library units.
    """
    
    def __init__(self, sync_report=False, intranet_sync_report=False, unit_report=False, **options):
        self.sync_report = sync_report
        self.intranet_sync_report = intranet_sync_report
        self.unit_report = unit_report
        self.options = {
            k: options[k] for k in (
                'all',
                'display_in_campus_directory',
                'latest_revision_created_at',
                'live'
            )
        }

    def workbook(self):
        """ 
        Returns:
            An OpenPyXL Workbook. 
        """
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        if self.unit_report:
            self._add_wagtail_units_report_worksheet()
        if self.sync_report:
            self._add_units_out_of_sync_worksheet()
        if self.intranet_sync_report:
            self._add_intranet_units_out_of_sync_worksheet()
        return self.workbook

    def tab_delimited(self):
        """
        Returns:
            A string. Tab-delimited fields with newlines between records.
        """
        output = ''
        if self.unit_report:
            output = output + self._get_wagtail_units_report_tab_delimited()
        if self.unit_report and self.sync_report:
            output = output + "\n\n\n"
        if self.sync_report:
            output = output + self._get_units_out_of_sync_tab_delimited()
        if self.intranet_sync_report:
            output = output + self._get_intranet_units_out_of_sync_tab_delimited()
        return output

    def _units_out_of_sync(self):
        """
        Get lists of unit pages that are out of sync.
    
        This function returns two lists--first, a list of the names of unit pages
        that are present in Wagtail, but missing in the campus directory. Second,
        a list of the names of unit pages that are present in the campus directory
        but missing in Wagtail.
    
        Note: We have to return names in the style of the campus directory (like
        those returned by the UnitPage's get_campus_directory_full_name() method)
        because we may only have access to a given unit in the campus directory. If
        that's the case it would be impossible to tell what its full name might be
        in the library directory.
    
        Returns: two lists of strings.
        """
        from units.models import UnitPage
        api_unit_names = self._get_campus_directory_unit_names()
        wag_unit_names = set()
    
        for u in UnitPage.objects.filter(
            display_in_campus_directory=True,
            live=True
        ):
            wag_unit_names.add(u.get_campus_directory_full_name())
    
        missing_in_campus_directory = sorted(
            list(api_unit_names.difference(wag_unit_names))
        )
        missing_in_wagtail = sorted(
            list(wag_unit_names.difference(api_unit_names))
        )
        return missing_in_campus_directory, missing_in_wagtail


    def _intranet_units_out_of_sync(self):
        """
        Get lists of intranet unit pages that are out of sync.
    
        This function returns two lists--first, a list of the names of intranet
        unit pages that are present in Wagtail, but missing in the campus
        directory.  Second, a list of the names of unit pages that are present
        in the campus directory but missing in Wagtail.
    
        Note: We have to return names in the style of the campus directory (like
        those returned by the UnitPage's get_campus_directory_full_name() method)
        because we may only have access to a given unit in the campus directory. If
        that's the case it would be impossible to tell what its full name might be
        in the library directory.
    
        Returns: two lists of strings.
        """
        from intranetunits.models import IntranetUnitsPage
        api_unit_names = self._get_campus_directory_unit_names()
        wag_unit_names = set()
    
        for u in IntranetUnitsPage.objects.live():
            wag_unit_names.add(u.get_campus_directory_full_name())
    
        missing_in_campus_directory = sorted(
            list(api_unit_names.difference(wag_unit_names))
        )
        missing_in_wagtail = sorted(
            list(wag_unit_names.difference(api_unit_names))
        )
        return missing_in_campus_directory, missing_in_wagtail


    def _get_campus_directory_unit_names(self):
        """
        Report unit names in the campus directory.
    
        Returns:
            a set() of campus directory full names, as strings.
        """
        unit_names = set()
        x = ElementTree.fromstring(
            get_xml_from_directory_api('https://directory.uchicago.edu/api/v2/divisions/16')
        )
        for d in x.findall(".//departments/department"):
            department_name = re.sub(
                '\s+',
                ' ',
                d.find('name').text
            ).strip()
            unit_names.add(department_name)
            department_xml = d.find('resources/xmlURL').text
            x2 = ElementTree.fromstring(
                get_xml_from_directory_api(department_xml)
            )
            for d2 in x2.findall(".//subDepartments/subDepartment"):
                subdepartment_name = re.sub(
                    '\s+',
                    ' ',
                    d2.find('name').text
                ).strip()
                unit_names.add(department_name + ' - ' + subdepartment_name)
        return unit_names

    def _get_units_out_of_sync_data(self):
        """
        Get the data for an out of sync units report, independant of whatever
        format the final report will be in (e.g. Excel or tab-delimited.)
        """
        output = []
        campus_units, wagtail_units = self._units_out_of_sync()
        if wagtail_units:
            output.append(["THE FOLLOWING UNITS APPEAR IN WAGTAIL, BUT NOT THE UNIVERSITY'S API:"])
            for w in wagtail_units:
                output.append([w])
            output.append([""])
        if campus_units:
            output.append(["THE FOLLOWING UNITS APPEAR IN THE UNIVERSITY'S API, BUT NOT WAGTAIL:"])
            for c in campus_units:
                output.append([c])
            output.append([""])
        return output

    def _get_intranet_units_out_of_sync_data(self):
        """
        Get the data for an out of sync units report, independant of whatever
        format the final report will be in (e.g. Excel or tab-delimited.)
        """
        output = []
        campus_units, wagtail_units = self._intranet_units_out_of_sync()
        if wagtail_units:
            output.append(["THE FOLLOWING UNITS APPEAR IN WAGTAIL, BUT NOT THE UNIVERSITY'S API:"])
            for w in wagtail_units:
                output.append([w])
            output.append([""])
        if campus_units:
            output.append(["THE FOLLOWING UNITS APPEAR IN THE UNIVERSITY'S API, BUT NOT WAGTAIL:"])
            for c in campus_units:
                output.append([c])
            output.append([""])
        return output

    def _add_units_out_of_sync_worksheet(self):
        """
        Adds a report of the units that are present in Wagtail but not in the
        campus directory, and vice versa, to a Microsoft Excel spreadsheet.
    
        Side Effect:
            Adds an OpenPyXL worksheet with information about out of sync units. 
        """
        worksheet = self.workbook.create_sheet('out of sync units')
        for record in self._get_units_out_of_sync_data():
            worksheet.append(record)

    def _add_intranet_units_out_of_sync_worksheet(self):
        """
        Adds a report of the units that are present in Wagtail but not in the
        campus directory, and vice versa, to a Microsoft Excel spreadsheet.
    
        Side Effect:
            Adds an OpenPyXL worksheet with information about out of sync units. 
        """
        worksheet = self.workbook.create_sheet('out of sync units')
        for record in self._get_intranet_units_out_of_sync_data():
            worksheet.append(record)

    def _get_units_out_of_sync_tab_delimited(self):
        """
        Get a report of library units that are present in Wagtail but not in
        the campus directory, and vice versa. 

        Returns:
            A string. Tab delimited data, separated by newlines. 
        """
        stringio = io.StringIO()
        writer = csv.writer(stringio, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
        for record in self._get_units_out_of_sync_data():
            writer.writerow(record)
        return stringio.getvalue()

    def _get_intranet_units_out_of_sync_tab_delimited(self):
        """
        Get a report of library units that are present in Wagtail but not in
        the campus directory, and vice versa. 

        Returns:
            A string. Tab delimited data, separated by newlines. 
        """
        stringio = io.StringIO()
        writer = csv.writer(stringio, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
        for record in self._get_intranet_units_out_of_sync_data():
            writer.writerow(record)
        return stringio.getvalue()

    def _get_units_wagtail(self):
        """
        Query for a list of UnitPage objects. The options passed to this function
        basically get applied as a Django filter().
    
        Returns:
            A sorted list of UnitPage objects.
        """
        from units.models import UnitPage
        try:
            if self.options['live']:
                unitpages = set(UnitPage.objects.live())
            else:
                unitpages = set(UnitPage.objects.all())
        except KeyError:
            unitpages = set(UnitPage.objects.all())
    
        try:
            if self.options['latest_revision_created_at']:
                latest_revision_created_at_string = '{}-{}-{} 00:00-0600'.format(
                    self.options['latest_revision_created_at'][0:4],
                    self.options['latest_revision_created_at'][4:6],
                    self.options['latest_revision_created_at'][6:8]
                )
                new_unitpages = set(UnitPage.objects.filter(latest_revision_created_at__gte=latest_revision_created_at_string))
                unitpages = unitpages.intersection(new_unitpages) if unitpages else new_unitpages
        except KeyError:
            pass
    
        try:
            if self.options['display_in_campus_directory']:
                new_unitpages = set(UnitPage.objects.filter(display_in_campus_directory=True))
                unitpages = unitpages.intersection(new_unitpages) if unitpages else new_unitpages
        except KeyError:
            pass
    
        return sorted(list(unitpages), key=lambda u: u.get_full_name())

    def _get_units_report_data(self):
        """
        Get the data for a report of library units, independant of whatever
        format the final report will be in (e.g. Excel or tab-delimited.)
        """
        output = []
        output.append([
            'ID',
            'LATEST REVISION CREATED AT',
            'LIBRARY DIRECTORY FULL NAME',
            'CAMPUS DIRECTORY FULL NAME'
        ])
        unitpages = self._get_units_wagtail()
        for u in unitpages:
            try:
                latest_revision_created_at = u.latest_revision_created_at.strftime('%m/%d/%Y %-I:%M:%S %p')
            except AttributeError:
                latest_revision_created_at = ''
            output.append([
                str(u.id),
                latest_revision_created_at,
                u.get_full_name(),
                u.get_campus_directory_full_name()
            ])
        return output

    def _add_wagtail_units_report_worksheet(self):
        """
        Add a report of library units in Microsoft Excel format, for HR reporting.
    
        Side Effect:
            Adds an OpenPyXL worksheet with information about units.
        """
        worksheet = self.workbook.create_sheet(title='wagtail units')
        for record in self._get_units_report_data():
            worksheet.append(record)

    def _get_wagtail_units_report_tab_delimited(self):
        """
        Get a report of library units in tab delimited format, for HR reporting.

        Returns:
            A string. Tab delimited data, separated by newlines. 
        """
        stringio = io.StringIO()
        writer = csv.writer(stringio, delimiter='\t', quoting=csv.QUOTE_MINIMAL)
        for record in self._get_units_report_data():
            writer.writerow(record)
        return stringio.getvalue()
