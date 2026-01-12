import os
from tempfile import NamedTemporaryFile

from diablo_tests import assert_assertion_error
from django.core import management
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.db.models.query import QuerySet
from django.test import Client, TestCase, override_settings
from openpyxl import load_workbook
from wagtail.models import Page, Site

from library_website.settings import PUBLIC_HOMEPAGE
from site_settings.models import QuickNumber, QuickNumberGroup
from staff.models import StaffPage

from .models import UnitPage
from .utils import (
    get_all_quick_nums_html,
    get_quick_num_html,
    get_quick_num_or_link,
    get_quick_nums_dict,
    get_quick_nums_for_library_or_dept,
)


class TestLibraryDirectory(TestCase):
    """
    Tests for the staff browse.
    """

    fixtures = ["test.json"]

    def test_subject_browse_dropdown(self):
        """
        Assert that the subjects context variable is always a
        QuerySet. A session must be saved in order to access
        context from the response.
        """
        client = Client()
        session = client.session
        session.save()
        response = client.get("/about/directory/?view=staff")
        qs = response.context["subjects"]
        is_qs = isinstance(qs, QuerySet)
        self.assertEqual(is_qs, True)


@override_settings(
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-quick-numbers",
        },
        "pagecache": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-pagecache",
        },
    }
)
class TestQuickNumberUtils(TestCase):
    """
    Tests for utilities used for adding quick numbers
    to the public staff directory.

    Uses local memory cache to prevent race conditions in parallel tests
    where multiple workers share Redis but have separate databases.
    """

    fixtures = ["test.json"]

    @classmethod
    def setUpTestData(cls):
        cls.quick_num = {"label": "Foobar", "number": "773-702-8740", "link": None}
        cls.quick_link = {"label": "Foobar", "number": "", "link": PUBLIC_HOMEPAGE}
        cls.dlist = [
            {"label": "Captain Janeway", "number": "00100", "link": None},
            {"label": "B'Elanna Torres", "number": "11000", "link": None},
            {"label": "Seven of Nine", "number": "01001", "link": None},
        ]
        cls.dlist_expected = "<td><strong>Captain Janeway</strong> 00100</td><td><strong>B'Elanna Torres</strong> 11000</td><td><strong>Seven of Nine</strong> 01001</td>"

    def test_get_quick_num_or_link(self):
        """
        Make sure the proper exceptions are raised when the input
        is bad and that the proper data is returned wheninput is good.
        """

        # For testing dict keys
        a = {"text": "Foobar", "number": "773-702-8740", "link": None}
        b = {"label": "Foobar", "link": None}

        # For testing dict key -> values
        c = {"label": "", "number": "773-702-8740", "link": None}
        d = {"label": "Foobar", "number": "773-702-8740", "link": "/"}

        # For testing return values
        e = {"label": "Foobar", "number": "773-702-8740", "link": None}
        f = {"label": "Foobar", "number": "", "link": PUBLIC_HOMEPAGE}
        g = {"label": "Foobar", "number": "", "link": None}

        # Test dict keys
        self.assertTrue(assert_assertion_error(get_quick_num_or_link, a))
        self.assertTrue(assert_assertion_error(get_quick_num_or_link, b))

        # Test dict key -> values
        self.assertTrue(assert_assertion_error(get_quick_num_or_link, c))
        self.assertTrue(assert_assertion_error(get_quick_num_or_link, d))

        # Test return values
        self.assertEqual(get_quick_num_or_link(e), (0, "Foobar", "773-702-8740"))
        self.assertEqual(get_quick_num_or_link(f), (1, "Foobar", "/"))
        self.assertRaises(ValueError, get_quick_num_or_link, g)

    def test_get_quick_num_html(self):
        """
        Test the expected HTML output.
        """
        # Expected HTML
        expected_num = "<td><strong>Foobar</strong>773-702-8740</td>"
        expected_link = '<td><strong><a href="/">Foobar</a></strong></td>'

        # Returned HTML
        quick_num_html = get_quick_num_html(get_quick_num_or_link(self.quick_num))
        quick_link_html = get_quick_num_html(get_quick_num_or_link(self.quick_link))

        # Test HTML
        self.assertHTMLEqual(quick_num_html, expected_num)
        self.assertHTMLEqual(quick_link_html, expected_link)

    def test_quick_num_base_config(self):
        """
        Test that quick numbers can be retrieved from Wagtail settings.
        No exceptions should be raised.
        """
        # Create test quick number group
        group = QuickNumberGroup.objects.create(
            slug="test-library", display_name="Test Library", is_default=True
        )
        QuickNumber.objects.create(
            group=group, label="Main Telephone", number="773-702-8740", link=None
        )

        quick_nums = get_quick_nums_dict()

        # Should get data from database
        self.assertIsInstance(quick_nums, dict)
        self.assertIn("test-library", quick_nums)

        # Test that each number dict has the required structure
        for library in quick_nums:
            for number in quick_nums[library]:
                get_quick_num_or_link(number)

    def test_quick_num_multiple_defaults_validation(self):
        """
        Test that only one QuickNumberGroup can be set as default.
        """
        # Create first default group
        group1 = QuickNumberGroup.objects.create(  # noqa: F841
            slug="library-one", display_name="Library One", is_default=True
        )

        # Try to create second default group - should fail validation
        group2 = QuickNumberGroup(
            slug="library-two", display_name="Library Two", is_default=True
        )

        with self.assertRaises(ValidationError):
            group2.full_clean()

    def test_get_quick_nums_for_library_or_dept_output(self):
        """
        Test the output of get_quick_nums_for_library_or_dept.
        """
        # Remove fixture's default group to avoid fallback interference
        QuickNumberGroup.objects.filter(is_default=True).delete()

        client = Client()
        site = Site.objects.filter(is_default_site=True)[0]

        # Create test quick number groups in database
        enterprise = QuickNumberGroup.objects.create(
            slug="enterprise", display_name="Enterprise"
        )
        QuickNumber.objects.create(
            group=enterprise, label="Captain Picard", number="11100", sort_order=0
        )
        QuickNumber.objects.create(
            group=enterprise, label="Worf Rozhenko", number="00001", sort_order=1
        )
        QuickNumber.objects.create(
            group=enterprise, label="Data", number="10111", sort_order=2
        )

        defiant = QuickNumberGroup.objects.create(
            slug="defiant", display_name="Defiant"
        )
        QuickNumber.objects.create(
            group=defiant, label="Benjamin Sisko", number="101001", sort_order=0
        )
        QuickNumber.objects.create(
            group=defiant, label="Quark", number="", link=site.root_page, sort_order=1
        )

        # Clear cache after creating test data to force fresh DB queries
        cache.clear()

        # Normal quick numbers
        expected = "<td><strong>Captain Picard</strong> 11100</td><td><strong>Worf Rozhenko</strong> 00001</td><td><strong>Data</strong> 10111</td>"
        response = client.get(
            "/about/directory/?view=staff&department=Enterprise",
            HTTP_HOST=site.hostname,
        )
        request = response.wsgi_request
        html = get_quick_nums_for_library_or_dept(request)
        self.assertHTMLEqual(html, expected)

        # Link instead of number
        expected = '<td><strong>Benjamin Sisko</strong> 101001</td><td><strong><a href="/">Quark</a></strong></td>'
        response = client.get(
            "/about/directory/?view=staff&department=Defiant",
            HTTP_HOST=site.hostname,
        )
        request = response.wsgi_request
        html = get_quick_nums_for_library_or_dept(request)
        self.assertHTMLEqual(html, expected)

        # Test library parameter
        voyager = QuickNumberGroup.objects.create(
            slug="voyager", display_name="Voyager"
        )
        for idx, item in enumerate(self.dlist):
            QuickNumber.objects.create(
                group=voyager,
                label=item["label"],
                number=item["number"],
                link=None,
                sort_order=idx,
            )

        expected = self.dlist_expected
        response = client.get(
            "/about/directory/?view=staff&library=Voyager", HTTP_HOST=site.hostname
        )
        request = response.wsgi_request
        html = get_quick_nums_for_library_or_dept(request)
        self.assertHTMLEqual(html, expected)

        # Bad parameter not in dictionary should return the default values
        # Clear any existing defaults from fixtures
        QuickNumberGroup.objects.filter(is_default=True).delete()

        default_group = QuickNumberGroup.objects.create(
            slug="test-default",
            display_name="Test Default Library",
            is_default=True,
        )
        QuickNumber.objects.create(
            group=default_group, label="Captain Long", number="00100", sort_order=0
        )
        QuickNumber.objects.create(
            group=default_group,
            label="Lieutenant Commander Blair",
            number="11000",
            sort_order=1,
        )

        expected = "<td><strong>Captain Long</strong> 00100</td><td><strong>Lieutenant Commander Blair</strong> 11000</td>"
        response = client.get(
            "/about/directory/?view=staff&department=Borg Cube",
            HTTP_HOST=site.hostname,
        )
        request = response.wsgi_request
        html = get_quick_nums_for_library_or_dept(request)
        self.assertHTMLEqual(html, expected)

    def test_get_all_quick_nums_html_output(self):
        """
        Test for expected html.
        """
        self.assertHTMLEqual(get_all_quick_nums_html(self.dlist), self.dlist_expected)


class ListUnitsWagtail(TestCase):

    def run_command(self, **options):
        tempfile = NamedTemporaryFile(delete=False, suffix=".xlsx")
        options.update({"filename": tempfile.name, "output_format": "excel"})
        management.call_command("list_units_wagtail", **options)

        wb = load_workbook(tempfile.name)
        ws = wb[wb.sheetnames[0]]
        os.unlink(tempfile.name)

        return [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]

    def setUp(self):
        try:
            welcome = Page.objects.get(path="00010001")
        except:  # noqa: E722
            root = Page.objects.create(depth=1, path="0001", slug="root", title="Root")

            welcome = Page(path="00010001", slug="welcome", title="Welcome")
            root.add_child(instance=welcome)

        staff_person = StaffPage(
            cnetid="staff-person", slug="staff-person", title="Staff Person"
        )
        welcome.add_child(instance=staff_person)

        some_unit = UnitPage(
            editor=staff_person,
            department_head=staff_person,
            page_maintainer=staff_person,
            slug="some-unit",
            title="Some Unit",
        )
        welcome.add_child(instance=some_unit)

    def test_report_column_count(self):
        records = self.run_command(all="True")
        self.assertEqual(len(records[0]), 4)

    def test_report_record_count(self):
        records = self.run_command(all="True")
        self.assertEqual(len(records), 1)
