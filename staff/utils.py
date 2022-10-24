# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import csv
import io
# import os
# os.environ['DJANGO_SETTINGS_MODULE'] = 'library_website.settings'
import re
from json import JSONDecodeError
from xml.etree import ElementTree

import requests

from base.utils import get_xml_from_directory_api, gensym
from django.contrib.auth.models import User
from library_website.settings import LIBCAL_TOKEN_ENDPOINT, LIBCAL_ENDPOINT, LIBCAL_CREDENTIALS, get_staff_url
from openpyxl import Workbook
from staff.models import EMPLOYEE_TYPES, StaffPage, StaffPageLibraryUnits
from units.models import UnitPage
from django.core.cache import cache

# need a list of all individuals.
# this thing needs to deal with VCards.


def get_all_library_cnetids_from_directory(xml_string=None):
    if not xml_string:
        xml_string = get_xml_from_directory_api(
            'https://directory.uchicago.edu/api/v2/divisions/16.xml'
        )

    # get xml element tree.
    x = ElementTree.fromstring(xml_string)

    # get cnetids.
    cnetids = set()
    for cnetid in x.findall(".//member/cnetid"):
        cnetids.add(cnetid.text)

    return sorted(list(cnetids))


def get_individual_info_from_directory(xml_string):
    x = ElementTree.fromstring(xml_string)

    info = {}

    info['cnetid'] = x.find("individuals/individual/cnetid").text

    # name is slightly more formal- e.g. "John E. Jung"
    info['officialName'] = x.find("individuals/individual/name").text

    # displayName is a bit more casual- e.g. "John Jung"
    info['displayName'] = x.find("individuals/individual/displayName").text

    #  title, department, subdepartment, building/room number, phone number:
    #  many individuals have more than one title- for example,
    #  Emily Treptow has the following two strings as titles:
    #  "Business and Economics Librarian for Instruction & Outreach"
    #  "Business & Economics Librarian for Instruction & Outreach"
    #  Show both titles, even though they're close, to encourage
    #  people to fix them in the university's system.
    #
    #  The title doesn't make sense without a department, subdepartment
    #  pair though. And many people, especially bibliographers,
    #  have three or four title/department/subdepartments.
    #
    # Also, some titles have different rooms- see Laura Ring for
    # an example. Many of those people, like Paul Belloni, will
    # have multiple phone numbers. (It looks like Paul has a phone
    # number at the SSA and at the Reg.)
    #
    # Each title gets it's own table? Each one points at a department/
    # subdepartment pair? (in it's own table...)

    info['title_department_subdepartments'] = set()
    info['title_department_subdepartments_dicts'] = []
    for vcard in x.findall("individuals/individual/contacts/contact"):
        try:
            if vcard.find('division/name').text != 'Library':
                continue
        except:
            continue

        output = []
        output_dict = {}

        try:
            title = re.sub('\s+', ' ', vcard.find('title').text).strip()
            if title:
                output.append(title)
                output_dict['title'] = title
        except:
            pass

        chunks = []
        try:
            department = re.sub('\s+', ' ',
                                vcard.find('department/name').text).strip()
            if department:
                chunks.append(department)
        except:
            pass

        try:
            subdepartment = re.sub(
                '\s+', ' ',
                vcard.find('subDepartment/name').text
            ).strip()
            if subdepartment:
                chunks = chunks + subdepartment.split(" | ")
        except:
            pass

        try:
            email = re.sub('\s+', ' ', vcard.find('email').text).strip()
            if email:
                output_dict['email'] = email
        except:
            pass

        try:
            facultyexchange = re.sub(
                '\s+', ' ',
                vcard.find('facultyExchange').text
            ).strip()
            if facultyexchange:
                output.append(facultyexchange)
                output_dict['facultyexchange'] = facultyexchange
        except:
            pass

        try:
            phone = re.sub('\s+', ' ', vcard.find('phone').text).strip()
            if phone:
                chunks = re.search(
                    '^\(([0-9]{3})\) ([0-9]{3})-([0-9]{4})$', phone
                )
                phone_number = chunks.group(1) + \
                    "-" + chunks.group(2) + "-" + chunks.group(3)
                output.append(phone_number)
                output_dict['phone'] = phone_number
        except:
            pass

        info['title_department_subdepartments'].add("\n".join(output))
        info['title_department_subdepartments_dicts'].append(output_dict)

    info['email'] = set()
    for vcard in x.findall("individuals/individual/contacts/contact"):
        try:
            if vcard.find('division/name').text != 'Library':
                continue
        except:
            continue

        try:
            email = re.sub('\s+', ' ', vcard.find('email').text).strip()
            if email:
                info['email'].add(email)
        except:
            pass

    info['phoneFacultyExchanges'] = set()
    info['phoneFacultyExchanges_dicts'] = []
    for vcard in x.findall("individuals/individual/contacts/contact"):
        try:
            if vcard.find('division/name').text != 'Library':
                continue
        except:
            continue

        output = []
        output_dict = {}

        try:
            facultyexchange = re.sub(
                '\s+', ' ',
                vcard.find('facultyExchange').text
            ).strip()
            if facultyexchange:
                output.append(facultyexchange)
                output_dict['facultyexchange'] = facultyexchange
        except:
            pass

        try:
            phone = re.sub('\s+', ' ', vcard.find('phone').text).strip()
            if phone:
                chunks = re.search(
                    '^\(([0-9]{3})\) ([0-9]{3})-([0-9]{4})$', phone
                )
                formatted_phone = chunks.group(1) + "-" + chunks.group(2) + \
                    "-" + chunks.group(3)
                output.append(formatted_phone)
                output_dict['phone'] = formatted_phone
        except:
            pass

        if output:
            info['phoneFacultyExchanges'].add("\n".join(output))
            info['phoneFacultyExchanges_dicts'].append(output_dict)

    info['positionTitle'] = ''
    for vcard in x.findall("individuals/individual/contacts/contact"):
        try:
            if vcard.find('division/name').text != 'Library':
                continue
        except:
            continue

        try:
            title = re.sub('\s+', ' ', vcard.find('title').text).strip()
            if title:
                info['positionTitle'] = title
                break
        except:
            pass

    info['departments'] = set()
    for vcard in x.findall("individuals/individual/contacts/contact"):
        try:
            if vcard.find('division/name').text != 'Library':
                continue
        except:
            continue

        department_name_pieces = []
        try:
            department_name_pieces.append(
                re.sub('\s+', ' ',
                       vcard.find('department/name').text.strip())
            )
        except:
            pass
        try:
            department_name_pieces.append(
                re.sub(
                    '\s+', ' ',
                    vcard.find('subDepartment/name').text.strip()
                )
            )
        except:
            pass
        info['departments'].add(' - '.join(department_name_pieces))

    return info


def get_all_library_cnetids_from_wagtail():
    output = []
    for s in StaffPage.objects.live():
        try:
            if User.objects.get(username=s.cnetid).is_active:
                output.append(s.cnetid)
        except:
            pass
    return output


def get_individual_info_from_wagtail(cnetid):
    staff_page = StaffPage.objects.get(cnetid=cnetid)

    if staff_page.display_name is None:
        raise ValueError(cnetid + ' has a display_name of None.')

    if staff_page.official_name is None:
        raise ValueError(cnetid + ' has an official_name of None.')

    # officialName is slightly more formal- e.g. "John E. Jung"
    # displayName is a bit more casual- e.g. "John Jung"
    output = {
        "cnetid": cnetid,
        "officialName": staff_page.official_name,
        "displayName": staff_page.display_name,
        "positionTitle": '',
        "email": set(),
        "departments": set(),
        "phoneFacultyExchanges": set()
    }

    for e in staff_page.staff_page_email.all():
        email_str = e.email.strip()
        if email_str:
            output['email'].add(email_str)

    if staff_page.position_title:
        output['positionTitle'] = re.sub('\s+', ' ',
                                         staff_page.position_title).strip()

    for v in staff_page.staff_page_phone_faculty_exchange.all():
        tmp = []

        faculty_exchange = v.faculty_exchange
        if faculty_exchange:
            tmp.append(re.sub('\s+', ' ', faculty_exchange).strip())

        phone_number = v.phone_number
        if phone_number:
            tmp.append(re.sub('\s+', ' ', phone_number).strip())

        if tmp:
            output['phoneFacultyExchanges'].add("\n".join(tmp))

    return output


class WagtailStaffReport:
    """
    Reporting on staff for the HR department. This class includes methods to
    report on the staff that are present in the campus directory but not in
    Wagtail, and it includes general reporting for library staff.
    """

    def __init__(self, sync_report=False, staff_report=False, **options):
        self.sync_report = sync_report
        self.staff_report = staff_report
        self.options = {
            k: options[k] for k in (
                'all', 'cnetid', 'department', 'department_and_subdepartments',
                'group', 'live', 'latest_revision_created_at',
                'position_eliminated', 'supervises_students',
                'supervisor_cnetid', 'supervisor_override', 'position_title'
            )
        }

    def _clean(self, string):
        """
        "Clean" a string that came in from the campus directory or from
        Wagtail. Remove leading spaces, replace multiple spaces with a single
        space, etc.
        """
        return re.sub('\s+', ' ', string).strip()

    def workbook(self):
        """
        Returns:
            An OpenPyXL Workbook.
        """
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        if self.staff_report:
            self._add_staff_report_worksheet()
        if self.sync_report:
            self._add_staff_out_of_sync_worksheet()
        return self.workbook

    def tab_delimited(self):
        """
        Returns:
            A string. Tab-delimited fields with newlines between records.
        """
        output = ''
        if self.staff_report:
            output = output + self._get_staff_report_tab_delimited()
        if self.staff_report and self.sync_report:
            output = output + "\n\n\n"
        if self.sync_report:
            output = output + self._get_staff_out_of_sync_tab_delimited()
        return output

    def _staff_out_of_sync(self):
        """
        Get lists of staff pages that are out of sync.

        This function returns two lists--first, a list of the names of staff
        pages that are present in Wagtail, but missing in the campus directory.
        Second, a list of the names of staff pages that are present in the
        campus directory but missing in Wagtail.

        Returns: two lists of strings.
        """

        def _format(cnetid, value, field):
            value = '' if value is None else str(value)
            return '{} -{}- ({})'.format(cnetid, self._clean(value), field)

        # Don't sync up some staff accounts. Former library directors may
        # appear in the campus directory, but they shouldn't appear in
        # staff listings on the library website. In other cases, staff
        # might have phone numbers connected with non-library jobs. If they
        # want those numbers to appear on the library website, they can add
        # them manually but we won't keep their information in sync.
        skip_cnetids = ['judi']

        api_staff_info = set()
        for cnetid in get_all_library_cnetids_from_directory():
            if cnetid in skip_cnetids:
                continue
            api_staff_info.add(cnetid)
            xml_string = get_xml_from_directory_api(
                'https://directory.uchicago.edu/api/v2/individuals/{}.xml'.
                format(cnetid)
            )
            info = get_individual_info_from_directory(xml_string)
            api_staff_info.add(
                _format(cnetid, info['officialName'], 'officialName')
            )
            api_staff_info.add(
                _format(cnetid, info['displayName'], 'displayName')
            )
            api_staff_info.add(
                _format(cnetid, info['positionTitle'], 'positionTitle')
            )
            for email in info['email']:
                api_staff_info.add(_format(cnetid, email, 'email'))
            for phone_facex in info['phoneFacultyExchanges']:
                api_staff_info.add(
                    _format(
                        cnetid, re.sub(r"\n", " ", phone_facex),
                        'phoneFacultyExchange'
                    )
                )
            for department in info['departments']:
                api_staff_info.add(_format(cnetid, department, 'department'))

        wag_staff_info = set()
        for s in StaffPage.objects.live():
            if s.cnetid in skip_cnetids:
                continue
            wag_staff_info.add(s.cnetid)
            wag_staff_info.add(
                _format(s.cnetid, s.official_name, 'officialName')
            )
            wag_staff_info.add(
                _format(s.cnetid, s.display_name, 'displayName'))
            wag_staff_info.add(
                _format(s.cnetid, s.position_title, 'positionTitle')
            )
            for e in s.staff_page_email.all():
                wag_staff_info.add(_format(s.cnetid, e.email, 'email'))
            for p in s.staff_page_phone_faculty_exchange.all():
                wag_staff_info.add(
                    _format(
                        s.cnetid,
                        '{} {}'.format(p.faculty_exchange,
                                       p.phone_number), 'phoneFacultyExchange'
                    )
                )
            for d in s.staff_page_units.all():
                wag_staff_info.add(
                    _format(
                        s.cnetid,
                        d.library_unit.get_campus_directory_full_name(),
                        'department'
                    )
                )

        missing_in_campus_directory = sorted(
            list(wag_staff_info.difference(api_staff_info))
        )
        missing_in_wagtail = sorted(
            list(api_staff_info.difference(wag_staff_info))
        )
        return missing_in_campus_directory, missing_in_wagtail

    def _get_staff_out_of_sync_data(self):
        """
        Get the data for an out of sync staff report, independant of whatever
        format the final report will be in (e.g. Excel or tab-delimited.)
        """
        output = []
        missing_in_campus_directory, missing_in_wagtail = self._staff_out_of_sync(
        )
        if missing_in_campus_directory:
            output.append(
                [
                    "THE FOLLOWING STAFF DATA APPEARS IN WAGTAIL, " +
                    "BUT NOT THE UNIVERSITY'S API:"
                ]
            )
            for c in missing_in_campus_directory:
                output.append([c])
            output.append([""])
        if missing_in_wagtail:
            output.append(
                [
                    "THE FOLLOWING STAFF DATA APPEARS IN THE " +
                    "UNIVERSITY'S API, BUT NOT WAGTAIL:"
                ]
            )
            for w in missing_in_wagtail:
                output.append([w])
            output.append([""])
        return output

    def _add_staff_out_of_sync_worksheet(self):
        """
        Adds a report of the staff that are present in Wagtail but not in the
        campus directory, and vice versa, to a Microsoft Excel spreadsheet.

        Side Effect:
            Adds an OpenPyXL worksheet with information about out of sync
            staff.
        """
        worksheet = self.workbook.create_sheet('out of sync staff')
        for record in self._get_staff_out_of_sync_data():
            worksheet.append(record)

    def _get_staff_out_of_sync_tab_delimited(self):
        """
        Get a report of library staff that are present in Wagtail but not in
        the campus directory, and vice versa.

        Returns:
            A string. Tab delimited data, separated by newlines.
        """
        stringio = io.StringIO()
        writer = csv.writer(stringio, delimiter='\t',
                            quoting=csv.QUOTE_MINIMAL)
        for record in self._get_staff_out_of_sync_data():
            writer.writerow(record)
        return stringio.getvalue()

    def _get_staff_wagtail(self):
        """
        Query for a list of UnitPage objects. The options passed to this
        function basically get applied as a Django filter().

        Returns:
            A sorted list of UnitPage objects.
        """
        try:
            if self.options['live']:
                staffpages = set(StaffPage.objects.live())
            else:
                staffpages = set(StaffPage.objects.all())
        except KeyError:
            staffpages = set(StaffPage.objects.all())

        filter_keys = ('cnetid', 'position_title')
        filter_options = {
            k: v for (k, v) in self.options.items() if k in filter_keys and v
        }

        if filter_options:
            new_staffpages = set(StaffPage.objects.filter(**filter_options))
            if staffpages:
                staffpages = staffpages.intersection(new_staffpages)
            else:
                staffpages = new_staffpages

        try:
            if self.options['supervises_students']:
                new_staffpages = set(
                    StaffPage.objects.filter(supervises_students=True)
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        try:
            if self.options['department']:
                library_units = [
                    u for u in UnitPage.objects.live()
                    if u.get_full_name() == self.options['department']
                ]
                new_staffpages = set(
                    StaffPage.objects.filter(
                        staff_page_units__library_unit__in=library_units
                    )
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        try:
            if self.options['department_and_subdepartments']:
                library_units = set()
                for u in [
                    u for u in UnitPage.objects.live() if u.get_full_name() ==
                    self.options['department_and_subdepartments']
                ]:
                    library_units = library_units.union(
                        set(u.get_descendants(True).type(UnitPage).specific())
                    )
                new_staffpages = set(
                    StaffPage.objects.filter(
                        staff_page_units__library_unit__in=list(library_units)
                    )
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        try:
            if self.options['group']:
                new_staffpages = set(
                    StaffPage.objects.filter(
                        member__parent__title=self.options['group']
                    )
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        try:
            if self.options['latest_revision_created_at']:
                l = '{}-{}-{} 00:00-0600'.format(
                    self.options['latest_revision_created_at'][0:4],
                    self.options['latest_revision_created_at'][4:6],
                    self.options['latest_revision_created_at'][6:8]
                )
                new_staffpages = set(
                    StaffPage.objects.filter(latest_revision_created_at__gte=l)
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        try:
            if self.options['position_eliminated']:
                new_staffpages = set(
                    StaffPage.objects.filter(position_eliminated=True)
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        try:
            if self.options['supervisor_override']:
                new_staffpages = set(
                    StaffPage.objects.exclude(supervisor_override=None)
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        try:
            if self.options['supervisor_cnetid']:
                new_staffpages = set(
                    StaffPage.objects.get(
                        cnetid=self.options['supervisor_cnetid']
                    ).get_staff()
                )
                if staffpages:
                    staffpages = staffpages.intersection(new_staffpages)
                else:
                    staffpages = new_staffpages
        except KeyError:
            pass

        return sorted(
            list(staffpages), key=lambda s: s.last_name if s.last_name else ''
        )

    def _get_staff_report_data(self):
        """
        Get the data for a report of library staff, independant of whatever
        format the final report will be in (e.g. Excel or tab-delimited.)
        """
        output = []
        output.append(
            [
                'ID', 'LATEST REVISION CREATED AT', 'NAME AND CNETID',
                'POSITION TITLE', 'EMAILS', 'PHONE, FACEX',
                'UNITS (LIBRARY DIRECTORY FULL NAME)',
                'UNITS (CAMPUS DIRECTORY FULL NAME)', 'GROUPS', 'EMPLOYEE TYPE',
                'SUPERVISES STUDENTS', 'POSITION ELIMINATED',
                'SUPERVISOR NAME AND CNETID'
            ]
        )
        staffpages = self._get_staff_wagtail()
        for s in staffpages:
            staffpage_library_units = set(
                StaffPageLibraryUnits.objects.filter(page=s)
            )

            units = []
            for slu in staffpage_library_units:
                units.append(
                    (
                        slu.library_unit.get_full_name(),
                        slu.library_unit.get_campus_directory_full_name()
                    )
                )
            units.sort(key=lambda u: u[0])

            groups = sorted([g.parent.title for g in s.member.all()])

            try:
                latest_revision_created_at = \
                    s.latest_revision_created_at.strftime(
                        '%m/%d/%Y %-I:%M:%S %p'
                    )
            except AttributeError:
                latest_revision_created_at = ''

            name_and_cnetid = '%s (%s)' % (s.title, s.cnetid)

            supervisor_names_and_cnetids = [
                '%s (%s)' % (s.title, s.cnetid) for s in s.get_supervisors
            ]

            emails = [e.email for e in s.staff_page_email.all()]

            phone_facexes = [
                '%s,%s' % (p.faculty_exchange, p.phone_number)
                for p in s.staff_page_phone_faculty_exchange.all()
            ]

            employee_type_string = [
                v for i, v in EMPLOYEE_TYPES if i == s.employee_type
            ][0]

            position_eliminated_string = str(s.position_eliminated)

            output.append(
                [
                    str(s.id), latest_revision_created_at, name_and_cnetid,
                    s.position_title or '', '|'.join(emails) or '',
                    '|'.join(phone_facexes) or '',
                    '|'.join([u[0] for u in units]) or '',
                    '|'.join([u[1] for u in units]) or '', '|'.join(groups),
                    employee_type_string,
                    str(s.supervises_students), position_eliminated_string,
                    '|'.join(supervisor_names_and_cnetids)
                ]
            )
        return output

    def _add_staff_report_worksheet(self):
        """
        Add a report of library staff in Microsoft Excel format, for HR
        reporting.

        Side Effect:
            Adds an OpenPyXL worksheet with information about staff.
        """
        worksheet = self.workbook.create_sheet(title='wagtail staff')
        for record in self._get_staff_report_data():
            worksheet.append(record)

    def _get_staff_report_tab_delimited(self):
        """
        Get a report of library staff in tab delimited format, for HR
        reporting.

        Returns:
            A string. Tab delimited data, separated by newlines.
        """
        stringio = io.StringIO()
        writer = csv.writer(stringio, delimiter='\t',
                            quoting=csv.QUOTE_MINIMAL)
        for record in self._get_staff_report_data():
            writer.writerow(record)
        return stringio.getvalue()


def get_token(url, data):
    """
    This function queries the LibCal API for an OAuth 2.0 token, as a
    first step to be used in lookup_staff_ids().

    Args:
        url string, dictionary representing POST headers

    Returns:
        access token string
    """
    try:
        resp = requests.post(url, data)
        return resp.json()['access_token']
    except(KeyError, requests.exceptions.RequestException):
        return ''


def lookup_staff_ids():
    """
    This function queries the LibCal API and returns a dictionary with
    email addresses as keys and LibCal IDs as values.

    Args:
        None

    Output:
        Email-to-LibCal ID lookup table
    """
    url = LIBCAL_ENDPOINT
    tok = get_token(LIBCAL_TOKEN_ENDPOINT, LIBCAL_CREDENTIALS)
    hdrs = {"Authorization": ("Bearer " + tok)}
    try:
        resp = requests.get(url, headers=hdrs)
        # the wrong URL will not return JSON
        json = resp.json()
        return {person['email']: person['id'] for person in json}
    except(JSONDecodeError, requests.exceptions.RequestException):
        return ''


def libcal_id_by_email(emailaddr):
    try:
        return lookup_staff_ids()[emailaddr]
    except (KeyError, TypeError):
        return ''



def staff_to_unit(s):
    stuff = s.staff_page_units.get_queryset()
    if len(stuff) > 0:
        return UnitPage.objects.get(pk=stuff.first().library_unit_id)
    else:
        return UnitPage.objects.get(title="Library")

 
def unit_to_staff(unit_id):
    return (StaffPage
            .objects
            .live()
            .filter(staff_page_units__library_unit_id=unit_id))

def make_staff_dict(staffpage):
    output = { "name" : staffpage.title,
               "node_name" : gensym(),
               "node_type" : "person",
               "url" : get_staff_url(staffpage) }
    return output

def make_org_dict(unit):
    non_draft_subunits = unit.get_children().live()
    def safe_head(unit):
        try:
            output = unit.specific.department_head.title
            head_url = get_staff_url(unit.specific.department_head)
        except AttributeError:
            output = "No Department Head Listed"
            head_url = ""
        return (output, head_url)
    if non_draft_subunits:
        subunit_json = [ make_org_dict(u) for u in non_draft_subunits ]
        node_type = "unit"
    else:
        staff = unit_to_staff(unit.id)
        subunit_json = [ make_staff_dict(s) for s in unit_to_staff(unit.id) ]
        node_type = "staff"
    output = { "head" : safe_head(unit)[0],
               "head_url" : safe_head(unit)[1],
               "name": unit.specific.title,
               "draft" : unit.has_unpublished_changes,
               "unit_id" : unit.id,
               "unit_url" : unit.url,
               "interim" : unit.specific.department_head_is_interim,
               "node_name" : gensym(),
               "node_type" : node_type,
               "subunits" : subunit_json }
    return output

def print_staff_dict(dct, tab_level=0):
    for person in dct["subunits"]:
        tab = "  " * tab_level
        print(tab + person["name"] + " (" + person["url"] + ")")

def print_org_dict(dct, tab_level=0):
    try:
        if dct["interim"]:
            head = (tab_level * "  ") + dct["head"] + " -- INTERIM --"
        else:
            head = (tab_level * "  ") + dct["head"]
        name = dct["name"]
        head_url = dct["head_url"]
        unit_url = dct["unit_url"]
        subs = dct["subunits"]
        print(head,
              "--",
              name,
              "(" + unit_url + ")")
        if dct["node_type"] == "staff":
            print_staff_dict(dct, tab_level=tab_level + 1)
        else:
            for s in subs:
                print_org_dict(s, tab_level=tab_level + 1)
    except KeyError:
        pass

# TODO: HTML code needs to be revised in light of changes to
# make_org_dict

# def entaggen(tag, text):
#     return "<%s>\n\t%s\n</%s>" % (tag, text, tag)

# def unordered_list(lst):
#     output = entaggen("ul", "".join([entaggen("li", x) for x in lst]))
#     return output

# def link_html(text, url):
#     parts = [
#         '<a href="',
#         url,
#         '">',
#         text,
#         "</a>",
#         ]
#     output = "".join(parts)
#     return output

# def head_link_html(dct):
#     text = dct["head"]
#     url = dct["head_url"]
#     return link_html(text, url)

# def staff_html(dct, head):
#     staff = unit_to_staff(dct["unit_id"])
#     non_draft_subunits = [
#         x for x in dct["subunits"] if not x["draft"]
#     ]
#     if non_draft_subunits == []:
#         list_elements = [ link_html(s.title,
#                                     get_staff_url(s))
#                           for s in staff
#                           if s.title != head ]
#         return unordered_list(list_elements)
#     else:
#         return ''

# def org_dict_to_html(dct):
#     if dct["interim"]:
#         interim = " -- INTERIM"
#     else:
#         interim = ""
#     name = entaggen("i", dct["name"])
#     unit_url = dct["unit_url"]
#     subs = dct["subunits"]
#     if not dct["draft"]:
#         the_rest = [org_dict_to_html(x) for x in subs if not x["draft"] ]
#         output = "%s %s -- %s%s" % (link_html(dct["head"], dct["head_url"]),
#                                     interim,
#                                     link_html(name, dct["unit_url"]),
#                                     unordered_list(the_rest))
#         staff_listing = staff_html(dct, dct["head"].strip())
#         if staff_listing:
#             return output + staff_listing
#         else:
#             return output
#     else:
#         return ''

def mk_graph(str):
    return "graph TD\n" + str

def trim_parens(str):
    return str.split("(")[0].strip()

def node_content(dct):
    if dct["node_type"] == "person":
        output = trim_parens(dct["name"])
    else:
        name = trim_parens(dct["name"])
        head = trim_parens(dct["head"])
        output = "%s<br>%s" % (name, head)
    return output

def org_chart_link(dct):
    format_string = "?view=org&unit=%i"
    return format_string % dct["unit_id"]

def unit_to_line(parent_dict, child_dict):
    parent_name = node_content(parent_dict)
    child_name = node_content(child_dict)
    format_string = ("%s[%s] --> %s[%s]\n"
                     "click %s \"%s\"\n"
                     "click %s \"%s\"\n")
    if parent_dict["node_type"] == "unit":
        url = org_chart_link(child_dict)
    else:
        url = child_dict["url"]
    return format_string % (parent_dict["node_name"],
                            parent_name,
                            child_dict["node_name"],
                            child_name,
                            parent_dict["node_name"],
                            org_chart_link(parent_dict),
                            child_dict["node_name"],
                            url,
                            )
def unit_to_lines(dct):
    current_daughters = "\n".join([ unit_to_line(dct, u)
                                    for u in dct["subunits"] ])
    if dct["node_type"] == "unit":
        recursive_daughters = "\n".join([ unit_to_lines(u)
                                          for u in dct["subunits"] ])
    else:
        recursive_daughters = ""
    return current_daughters + recursive_daughters

def org_dict_to_mermaid(dct):
    return mk_graph(unit_to_lines(dct))

def cache_unit_json(unit):
    cache_key = "org_chart_" + str(unit.id)
    cache.set(cache_key, make_org_dict(unit))

def cache_lookup(unit):
    cache_key = "org_chart_" + str(unit.id)
    return cache.get(cache_key)

def update_org_chart_cache():
    for u in UnitPage.objects.live():
        cache_unit_json(unit)

def depth(dct):
    if dct["node_type"] != "person":
        if dct["subunits"]:
            return 1 + max([ depth(u) for u in dct["subunits"] ])
        else:
            return 0
    else:
        return 0

def trim(dct, dpth):
    if dct["node_type"] == "unit":
        output
    # if dpth == 0:
    #     dct["subunits"] = []
    # else:
    #     if dct["node_type"] != "person":
    #         dct["subunits"] = [ trim(u, dpth - 1) for u in dct["subunits"] ]
    #     else:
    #         pass
        
    
    pass
