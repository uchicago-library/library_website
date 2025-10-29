import os
from tempfile import NamedTemporaryFile

from django.contrib.auth.models import Group, Permission, User
from django.core import management
from django.test import TestCase
from openpyxl import load_workbook
from public.models import StaffPublicPage, StandardPage
from units.models import UnitPage
from wagtail.models import Page

from staff.utils import (
    get_all_library_cnetids_from_directory,
    get_individual_info_from_directory,
)

from .models import (
    StaffPage,
    StaffPageEmailAddresses,
    StaffPageLibraryUnits,
    StaffPagePhoneFacultyExchange,
)


class UniversityDirectoryTestCase(TestCase):
    def test_get_all_library_cnet_ids_from_directory_two_staff(self):
        xml_string = """<?xml version="1.0" encoding="UTF-8"?>
        <responseData>
          <response>Success</response>
          <totalResults>1</totalResults>
          <organizations>
            <organization>
              <name>Library</name>
              <type>Division</type>
              <departments>
                <department>
                  <name>Digital Services</name>
                  <resources>
                    <directoryURL>https://directory.uchicago.edu/organizations/575?type=departments</directoryURL>
                    <xmlURL>https://directory.uchicago.edu/api/v2/departments/575</xmlURL>
                  </resources>
                </department>
              </departments>
              <members>
                <member>
                  <name>Scooter McDanger</name>
                  <displayName>Scooter McDanger</displayName>
                  <cnetid>danger</cnetid>
                  <chicagoid>12345678L</chicagoid>
                  <title>Programmer/Analyst</title>
                  <email>danger@uchicago.edu</email>
                  <phone>(773) 702-1234</phone>
                  <facultyExchange>JRL 220</facultyExchange>
                  <resources>
                    <directoryURL>https://directory.uchicago.edu/individuals/12345678L</directoryURL>
                    <xmlURL>https://directory.uchicago.edu/api/v2/individuals/12345678L</xmlURL>
                  </resources>
                </member>
                <member>
                  <name>Margaret Lovelee</name>
                  <displayName>Margaret Lovelee</displayName>
                  <cnetid>lovelee</cnetid>
                  <chicagoid>12345678L</chicagoid>
                  <title>Programmer/Analyst</title>
                  <email>lovelee@uchicago.edu</email>
                  <phone>(773) 702-1234</phone>
                  <facultyExchange>JRL 220</facultyExchange>
                  <resources>
                    <directoryURL>https://directory.uchicago.edu/individuals/12345678L</directoryURL>
                    <xmlURL>https://directory.uchicago.edu/api/v2/individuals/12345678L</xmlURL>
                  </resources>
                </member>
              </members>
              <resources>
                <directoryURL>https://directory.uchicago.edu/organizations/16?type=divisions</directoryURL>
                <xmlURL>https://directory.uchicago.edu/api/v2/divisions/16</xmlURL>
              </resources>
            </organization>
          </organizations>
        </responseData>
        """

        cnetids = get_all_library_cnetids_from_directory(xml_string)

        self.assertEqual(set(cnetids), set(('lovelee', 'danger')))

    def assertInfoEqual(self, a, b):
        for field in ['cnetid', 'officialName', 'displayName']:
            if not a[field] == b[field]:
                return False

        tds_a = sorted([a['title_department_subdepartments']])
        tds_b = sorted([b['title_department_subdepartments']])

        if not len(tds_a) == len(tds_b):
            return False

        t = 0
        while t < len(tds_a):
            if not tds_a[t] == tds_b[t]:
                return False
            t = t + 1

        tds_a = sorted(
            a['title_department_subdepartments_dicts'],
            key=lambda t: t['title'] + t['department'],
        )
        tds_b = sorted(
            b['title_department_subdepartments_dicts'],
            key=lambda t: t['title'] + t['department'],
        )

        if not len(tds_a) == len(tds_b):
            return False

        t = 0
        while t < len(tds_a):
            for field in ['department', 'title', 'email', 'facultyexchange', 'phone']:
                if not tds_a[t][field] == tds_b[t][field]:
                    return False
            t = t + 1

        return True

    def test_get_individual_info_from_directory(self):
        self.maxDiff = None

        xml = """<?xml version="1.0" encoding="UTF-8"?>
        <responseData>
          <response>Success</response>
          <totalResults>1</totalResults>
          <individuals>
            <individual>
              <name>Scooter McDanger</name>
              <displayName>Scooter McDanger</displayName>
              <cnetid>scootermcdanger</cnetid>
              <chicagoid>12345678X</chicagoid>
              <contacts>
                <contact>
                  <title>Programmer Analyst</title>
                  <division>
                    <name>Library</name>
                    <resources>
                      <directoryURL>https://directory.uchicago.edu/organizations/16?type=divisions</directoryURL>
                      <xmlURL>https://directory.uchicago.edu/api/v2/divisions/16</xmlURL>
                    </resources>
                  </division>
                  <email>scootermcdanger@uchicago.edu</email>
                  <phone>(773) 702-1234</phone>
                  <facultyExchange>JRL 220</facultyExchange>
                </contact>
              </contacts>
              <resources>
                <directoryURL>https://directory.uchicago.edu/individuals/12345678X</directoryURL>
                <xmlURL>https://directory.uchicago.edu/api/v2/individuals/12345678X</xmlURL>
              </resources>
            </individual>
          </individuals>
        </responseData>
        """

        info = {
            'cnetid': 'scootermcdanger',
            'officialName': 'Scooter McDanger',
            'displayName': 'Scooter McDanger',
            'title_department_subdepartments': {
                'Programmer Analyst\nLibrary\nJRL 220\n773-702-1234'
            },
            'title_department_subdepartments_dicts': [
                {
                    'title': 'Programmer Analyst',
                    'email': 'scootermcdanger@uchicago.edu',
                    'facultyexchange': 'JRL 220',
                    'phone': '773-702-1234',
                }
            ],
        }

        self.assertInfoEqual(get_individual_info_from_directory(xml), info)


class StaffPageSupervisors(TestCase):
    @classmethod
    def setUpTestData(cls):
        # Create page hierarchy (runs once for all tests in this class)
        try:
            welcome = Page.objects.get(path='00010001')
        except:
            root = Page.objects.create(depth=1, path='0001', slug='root', title='Root')

            welcome = Page(path='00010001', slug='welcome', title='Welcome')
            root.add_child(instance=welcome)

        # StaffPages
        official_supervisor = StaffPage(
            cnetid='official_supervisor',
            slug='official-supervisor',
            title='Official Supervisor',
        )
        welcome.add_child(instance=official_supervisor)

        another_official_supervisor = StaffPage(
            cnetid='another_official_supervisor',
            slug='another-official-supervisor',
            title='Another Official Supervisor',
        )
        welcome.add_child(instance=another_official_supervisor)

        director = StaffPage(cnetid='director', slug='director', title='Director')
        welcome.add_child(instance=director)

        supervisor_override = StaffPage(
            cnetid='supervisor_override',
            slug='supervisor-override',
            title='Supervisor Override',
        )
        welcome.add_child(instance=supervisor_override)

        employee_override = StaffPage(
            cnetid='employee_override',
            slug='employee_override',
            supervisor_override=supervisor_override,
            title='Employee Override',
        )
        welcome.add_child(instance=employee_override)

        employee_no_override = StaffPage(
            cnetid='employee_no_override',
            slug='employee_no_override',
            title='Employee No Override',
        )
        welcome.add_child(instance=employee_no_override)

        employee_two_units = StaffPage(
            cnetid='employee_two_units',
            slug='employee_two_units',
            title='Employee Two Units',
        )
        welcome.add_child(instance=employee_two_units)

        content_specialist = StaffPage(
            cnetid='content_specialist',
            slug='content-specialist',
            title='Content Specialist',
        )
        welcome.add_child(instance=content_specialist)

        editor = StaffPage(cnetid='editor', slug='editor', title='Editor')
        welcome.add_child(instance=editor)

        page_maintainer = StaffPage(
            cnetid='page_maintainer', slug='page-maintainer', title='Page Maintainer'
        )
        welcome.add_child(instance=page_maintainer)

        # UnitPages
        unit_division = UnitPage(
            department_head=director,
            editor=editor,
            page_maintainer=page_maintainer,
            slug='unit-division',
            title='Unit Division',
        )
        welcome.add_child(instance=unit_division)

        unit_one = UnitPage(
            department_head=official_supervisor,
            display_in_dropdown=True,
            editor=editor,
            page_maintainer=page_maintainer,
            slug='unit-one',
            title='Unit One',
        )
        unit_division.add_child(instance=unit_one)

        unit_two = UnitPage(
            department_head=another_official_supervisor,
            editor=editor,
            page_maintainer=page_maintainer,
            slug='unit-two',
            title='Unit Two',
        )
        unit_division.add_child(instance=unit_two)

        # StandardPage
        standard_page = StandardPage(
            content_specialist=content_specialist,
            editor=editor,
            page_maintainer=page_maintainer,
            slug='standard-page',
            title='Standard Page',
            unit=unit_one,
        )
        welcome.add_child(instance=standard_page)

        # assign staff to units
        StaffPageLibraryUnits.objects.create(page=director, library_unit=unit_division)

        StaffPageLibraryUnits.objects.create(
            page=official_supervisor, library_unit=unit_one
        )

        StaffPageLibraryUnits.objects.create(
            page=employee_no_override, library_unit=unit_one
        )

        StaffPageLibraryUnits.objects.create(
            page=employee_override, library_unit=unit_one
        )

        StaffPageLibraryUnits.objects.create(
            page=employee_two_units, library_unit=unit_one
        )

        StaffPageLibraryUnits.objects.create(
            page=employee_two_units, library_unit=unit_two
        )

    def test_supervisor_relationships(self):
        # official supervisor
        self.assertEqual(
            StaffPage.objects.get(cnetid='employee_no_override').get_supervisors,
            [StaffPage.objects.get(cnetid='official_supervisor')],
        )

        # supervisor override
        self.assertEqual(
            StaffPage.objects.get(cnetid='employee_override').get_supervisors,
            [StaffPage.objects.get(cnetid='supervisor_override')],
        )

        # employee in two units
        self.assertEqual(
            StaffPage.objects.get(cnetid='employee_two_units').get_supervisors,
            [
                StaffPage.objects.get(cnetid='official_supervisor'),
                StaffPage.objects.get(cnetid='another_official_supervisor'),
            ],
        )

        # department head supervisor
        self.assertEqual(
            StaffPage.objects.get(cnetid='official_supervisor').get_supervisors,
            [StaffPage.objects.get(cnetid='director')],
        )

        # one child of division.
        self.assertEqual(
            UnitPage.objects.get(slug='unit-one').get_parent().specific,
            UnitPage.objects.get(slug='unit-division'),
        )


class ListStaffWagtail(TestCase):
    @classmethod
    def setUpTestData(cls):
        # Create page hierarchy (runs once for all tests in this class)
        try:
            welcome = Page.objects.get(path='00010001')
        except:
            root = Page.objects.create(depth=1, path='0001', slug='root', title='Root')

            welcome = Page(path='00010001', slug='welcome', title='Welcome')
            root.add_child(instance=welcome)

        chas = StaffPage(
            cnetid='chas',
            employee_type=3,
            slug='charles-blair',
            position_eliminated=False,
            position_title='Director, Digital Library Development Center',
            supervises_students=True,
            title='Charles Blair',
        )
        welcome.add_child(instance=chas)

        bbusenius = StaffPage(
            cnetid='bbusenius',
            employee_type=3,
            slug='brad-busenius',
            position_eliminated=False,
            position_title='Web Administrator',
            supervisor_override=chas,
            supervises_students=False,
            title='Brad Busenius',
        )
        welcome.add_child(instance=bbusenius)

        byrne = StaffPage(
            cnetid='byrne',
            employee_type=3,
            slug='maura-byrne',
            position_eliminated=False,
            position_title='Applications Systems Analyst/Programmer',
            supervises_students=False,
            title='Maura Byrne',
        )
        welcome.add_child(instance=byrne)

        eliminated_position = StaffPage(
            cnetid='eliminated-position',
            slug='eliminated-position',
            position_eliminated=False,
            title='Eliminated Position',
        )
        welcome.add_child(instance=eliminated_position)

        elong = StaffPage(
            cnetid='elong',
            slug='elisabeth-long',
            position_eliminated=False,
            title='Elisabeth Long',
        )
        welcome.add_child(instance=elong)

        digital_services = UnitPage(
            department_head=elong,
            editor=elong,
            page_maintainer=elong,
            slug='digital-services',
            title='Digital Services',
        )
        welcome.add_child(instance=digital_services)

        dldc = UnitPage(
            department_head=chas,
            editor=chas,
            page_maintainer=chas,
            slug='dldc',
            title='Digital Library Development Center',
        )
        digital_services.add_child(instance=dldc)

        jej = StaffPage(
            cnetid='jej',
            employee_type=3,
            slug='john-jung',
            position_eliminated=False,
            position_title='Programmer/Analyst',
            supervisor_override=chas,
            supervises_students=False,
            title='John Jung',
        )
        welcome.add_child(instance=jej)

        StaffPageEmailAddresses.objects.create(page=jej, email='jej@uchicago.edu')

        StaffPageEmailAddresses.objects.create(page=jej, email='jej@jej.com')

        StaffPagePhoneFacultyExchange.objects.create(
            page=jej, phone_number='773-702-1234', faculty_exchange='JRL 100'
        )

        StaffPagePhoneFacultyExchange.objects.create(
            page=jej, phone_number='773-834-1234', faculty_exchange='JRL 101'
        )

        StaffPageLibraryUnits.objects.create(page=jej, library_unit=dldc)

        StaffPageLibraryUnits.objects.create(page=jej, library_unit=digital_services)

        kzadrozny = StaffPage(
            cnetid='kzadrozny',
            employee_type=3,
            slug='kathy-zadrozny',
            position_eliminated=False,
            position_title='Web Developer and Graphic Design Specialist',
            supervisor_override=chas,
            supervises_students=True,
            title='Kathy Zadrozny',
        )
        welcome.add_child(instance=kzadrozny)

        tyler = StaffPage(
            cnetid='tyler',
            employee_type=3,
            slug='tyler-danstrom',
            position_eliminated=False,
            position_title='Programmer/Analyst',
            supervisor_override=chas,
            supervises_students=False,
            title='Tyler Danstrom',
        )
        welcome.add_child(instance=tyler)

    def run_command(self, **options):
        tempfile = NamedTemporaryFile(delete=False, suffix='.xlsx')
        options.update({'filename': tempfile.name, 'output_format': 'excel'})
        management.call_command('list_staff_wagtail', **options)

        wb = load_workbook(tempfile.name)
        ws = wb.active
        os.unlink(tempfile.name)

        return [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]

    def test_report_columns(self):
        records = self.run_command(cnetid='jej')

        # column count
        self.assertEqual(len(records[0]), 13)

        # row count
        self.assertEqual(len(records), 1)

        # name and cnetid
        self.assertEqual(records[0][2], 'John Jung (jej)')

        # position title
        self.assertEqual(records[0][3], 'Programmer/Analyst')

        # emails
        self.assertEqual(
            set(records[0][4].split('|')), set(('jej@uchicago.edu', 'jej@jej.com'))
        )

        # faculty exchange, phone number pairs
        self.assertEqual(
            set(records[0][5].split('|')),
            set(('JRL 100,773-702-1234', 'JRL 101,773-834-1234')),
        )

        # units
        self.assertEqual(
            set(records[0][6].split('|')),
            set(
                [
                    'Digital Services - Digital Library Development Center',
                    'Digital Services',
                ]
            ),
        )

        # employee type
        self.assertEqual(records[0][9], 'IT')

        # supervises students
        self.assertEqual(records[0][10], 'False')

        # position eliminated
        self.assertEqual(records[0][11], 'False')

        # supervisor
        self.assertEqual(records[0][12].rstrip(), 'Charles Blair (chas)')

    def test_report_queries(self):
        # position title
        records = self.run_command(position_title='Programmer/Analyst')
        self.assertEqual(len(records), 2)

        # supervisor
        records = self.run_command(supervisor_cnetid='chas')
        self.assertEqual(len(records), 4)

        # supervises students
        records = self.run_command(supervises_students=True)
        self.assertEqual(len(records), 2)

    def test_empty_library_unit_does_not_crash(self):
        """
        Test that staff with empty library units (None) don't cause crashes
        in the reporting commands. Regression test for issue #702.
        """
        # Get a staff member
        staff = StaffPage.objects.get(cnetid='jej')

        # Add an empty library unit (None) - this reproduces the bug
        StaffPageLibraryUnits.objects.create(page=staff, library_unit=None)

        # The staff report should not crash
        records = self.run_command(cnetid='jej')
        self.assertEqual(len(records), 1)

        # Out of sync report should also not crash
        tempfile = NamedTemporaryFile(delete=False, suffix='.xlsx')
        management.call_command(
            'list_staff_wagtail',
            filename=tempfile.name,
            output_format='excel',
            report_out_of_sync_staff=True
        )

        wb = load_workbook(tempfile.name)
        # Should have at least one worksheet
        self.assertGreaterEqual(len(wb.sheetnames), 1)
        os.unlink(tempfile.name)


class StaffPageHRPermissionsTestCase(TestCase):
    """
    Test suite for Human Resources Info tab permissions.
    Verifies that the custom permission restricts access to HR fields
    on StaffPages as intended.
    """

    def setUp(self):
        """Set up test users, groups, and a staff page."""
        # Create root and welcome pages
        try:
            welcome = Page.objects.get(path='00010001')
        except Page.DoesNotExist:
            root = Page.objects.create(depth=1, path='0001', slug='root', title='Root')
            welcome = Page(path='00010001', slug='welcome', title='Welcome')
            root.add_child(instance=welcome)

        # Create a test StaffPage
        self.staff_page = StaffPage(
            cnetid='testuser',
            slug='test-user',
            title='Test User',
            position_title='Test Position',
        )
        welcome.add_child(instance=self.staff_page)

        # Create test users
        self.regular_user = User.objects.create_user(
            username='regular', password='password'
        )

        self.hr_user = User.objects.create_user(username='hruser', password='password')

        self.superuser = User.objects.create_superuser(
            username='admin', password='password', email='admin@example.com'
        )

        # Create HR group with the custom permission
        self.hr_group = Group.objects.create(name='HR Staff')
        self.hr_permission = Permission.objects.get(
            codename='change_staff_hr_info', content_type__app_label='staff'
        )
        self.hr_group.permissions.add(self.hr_permission)
        self.hr_user.groups.add(self.hr_group)

    def test_hr_permission_exists(self):
        """Test that the custom HR permission was created."""
        permission = Permission.objects.filter(
            codename='change_staff_hr_info', content_type__app_label='staff'
        )
        self.assertEqual(permission.count(), 1)
        self.assertEqual(
            permission.first().name, 'Can edit Human Resources Info for staff pages'
        )

    def test_regular_user_lacks_hr_permission(self):
        """Test that regular users don't have HR permission."""
        self.assertFalse(self.regular_user.has_perm('staff.change_staff_hr_info'))

    def test_hr_user_has_hr_permission(self):
        """Test that HR users have the custom permission."""
        self.assertTrue(self.hr_user.has_perm('staff.change_staff_hr_info'))

    def test_superuser_has_hr_permission(self):
        """Test that superusers have the HR permission."""
        self.assertTrue(self.superuser.has_perm('staff.change_staff_hr_info'))

    def test_hr_permission_in_group(self):
        """Test that the permission can be assigned to groups."""
        self.assertIn(self.hr_permission, self.hr_group.permissions.all())

    def test_staffpage_has_permission_in_meta(self):
        """Test that StaffPage model defines the custom permission."""
        self.assertIn(
            ('change_staff_hr_info', 'Can edit Human Resources Info for staff pages'),
            StaffPage._meta.permissions,
        )


class UserActiveStateStaffPagePublishingTestCase(TestCase):
    """
    Test suite for automatic staff page publishing/unpublishing
    when user active state changes.
    """

    @classmethod
    def setUpTestData(cls):
        """Set up test users and staff pages."""
        # Create root and welcome pages
        try:
            welcome = Page.objects.get(path='00010001')
        except Page.DoesNotExist:
            root = Page.objects.create(depth=1, path='0001', slug='root', title='Root')
            welcome = Page(path='00010001', slug='welcome', title='Welcome')
            root.add_child(instance=welcome)

        cls.welcome = welcome

        # Create a staff page to use as editor/maintainer/content specialist
        admin_staff = StaffPage(
            cnetid='admin',
            slug='admin-staff',
            title='Admin Staff',
        )
        welcome.add_child(instance=admin_staff)

        # Create a unit page for required unit field
        cls.unit = UnitPage(
            slug='test-unit',
            title='Test Unit',
            display_in_dropdown=True,
            editor=admin_staff,
            page_maintainer=admin_staff,
        )
        welcome.add_child(instance=cls.unit)

        cls.admin_staff = admin_staff

    def setUp(self):
        """Set up per-test data."""
        # Create test user
        self.user = User.objects.create_user(
            username='teststaff', password='password', is_active=True
        )

        # Create Loop staff page
        self.staff_page = StaffPage(
            cnetid='teststaff',
            slug='test-staff',
            title='Test Staff',
            position_title='Test Position',
        )
        self.welcome.add_child(instance=self.staff_page)
        self.staff_page.save_revision().publish()

        # Create public staff page
        self.public_page = StaffPublicPage(
            cnetid='teststaff',
            slug='test-staff-public',
            title='Test Staff Public',
            editor=self.admin_staff,
            page_maintainer=self.admin_staff,
            content_specialist=self.admin_staff,
            unit=self.unit,
        )
        self.welcome.add_child(instance=self.public_page)
        self.public_page.save_revision().publish()

    def test_deactivating_user_unpublishes_staff_pages(self):
        """Test that marking a user inactive unpublishes their staff pages."""
        # Verify pages are published
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertTrue(self.staff_page.live)
        self.assertTrue(self.public_page.live)

        # Deactivate user
        self.user.is_active = False
        self.user.save()

        # Verify pages are unpublished
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertFalse(self.staff_page.live)
        self.assertFalse(self.public_page.live)

    def test_reactivating_user_republishes_staff_pages(self):
        """Test that marking an inactive user active republishes their staff pages."""
        # Deactivate user first
        self.user.is_active = False
        self.user.save()

        # Verify pages are unpublished
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertFalse(self.staff_page.live)
        self.assertFalse(self.public_page.live)

        # Reactivate user
        self.user.is_active = True
        self.user.save()

        # Verify pages are republished
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertTrue(self.staff_page.live)
        self.assertTrue(self.public_page.live)

    def test_activating_inactive_user_with_unpublished_pages(self):
        """Test that activating an existing inactive user publishes their unpublished pages."""
        # Create an inactive user
        inactive_user = User.objects.create_user(
            username='inactiveuser', password='password', is_active=False
        )

        # Create unpublished staff pages for this user
        unpublished_staff = StaffPage(
            cnetid='inactiveuser',
            slug='inactive-user-staff',
            title='Inactive User Staff',
        )
        self.welcome.add_child(instance=unpublished_staff)
        unpublished_staff.unpublish()

        unpublished_public = StaffPublicPage(
            cnetid='inactiveuser',
            slug='inactive-user-public',
            title='Inactive User Public',
            editor=self.admin_staff,
            page_maintainer=self.admin_staff,
            content_specialist=self.admin_staff,
            unit=self.unit,
        )
        self.welcome.add_child(instance=unpublished_public)
        unpublished_public.unpublish()

        # Verify pages are unpublished
        unpublished_staff.refresh_from_db()
        unpublished_public.refresh_from_db()
        self.assertFalse(unpublished_staff.live)
        self.assertFalse(unpublished_public.live)

        # Activate the user
        inactive_user.is_active = True
        inactive_user.save()

        # Verify pages are now published
        unpublished_staff.refresh_from_db()
        unpublished_public.refresh_from_db()
        self.assertTrue(unpublished_staff.live)
        self.assertTrue(unpublished_public.live)

    def test_deactivating_user_with_no_staff_pages(self):
        """Test that deactivating a user with no staff pages doesn't raise errors."""
        # Create user without staff pages
        user_no_pages = User.objects.create_user(
            username='nostaff', password='password', is_active=True
        )

        # This should not raise an exception
        user_no_pages.is_active = False
        user_no_pages.save()

        # Verify user is inactive
        user_no_pages.refresh_from_db()
        self.assertFalse(user_no_pages.is_active)

    def test_creating_new_inactive_user_does_not_affect_existing_published_pages(self):
        """Test that creating a new inactive user matching existing published pages doesn't unpublish them."""
        # Verify the existing pages are published
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertTrue(self.staff_page.live)
        self.assertTrue(self.public_page.live)

        # Create a new inactive user with matching cnetid (teststaff)
        # This should NOT unpublish the existing pages because the signal only fires
        # when is_active changes from True to False, not when a new user is created as inactive
        User.objects.create_user(
            username='teststaff2', password='password', is_active=False
        )

        # Create additional published staff pages for this new user
        new_staff_page = StaffPage(
            cnetid='teststaff2',
            slug='test-staff-2',
            title='Test Staff 2',
        )
        self.welcome.add_child(instance=new_staff_page)
        new_staff_page.save_revision().publish()

        new_public_page = StaffPublicPage(
            cnetid='teststaff2',
            slug='test-staff-2-public',
            title='Test Staff 2 Public',
            editor=self.admin_staff,
            page_maintainer=self.admin_staff,
            content_specialist=self.admin_staff,
            unit=self.unit,
        )
        self.welcome.add_child(instance=new_public_page)
        new_public_page.save_revision().publish()

        # The new pages should remain published even though the user was created as inactive
        # because the signal doesn't fire on user creation, only on save when is_active changes
        new_staff_page.refresh_from_db()
        new_public_page.refresh_from_db()
        self.assertTrue(new_staff_page.live)
        self.assertTrue(new_public_page.live)

        # Original pages should still be published
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertTrue(self.staff_page.live)
        self.assertTrue(self.public_page.live)

    def test_creating_new_active_user_publishes_matching_unpublished_pages(self):
        """Test that creating a new active user publishes their unpublished staff pages."""
        # Create unpublished staff pages with a cnetid that doesn't have a user yet
        unpublished_staff = StaffPage(
            cnetid='newactive', slug='new-active-staff', title='New Active Staff'
        )
        self.welcome.add_child(instance=unpublished_staff)
        unpublished_staff.unpublish()

        # Create an unpublished public page too
        unpublished_public = StaffPublicPage(
            cnetid='newactive',
            slug='new-active-public',
            title='New Active Public',
            editor=self.admin_staff,
            page_maintainer=self.admin_staff,
            content_specialist=self.admin_staff,
            unit=self.unit,
        )
        self.welcome.add_child(instance=unpublished_public)
        unpublished_public.unpublish()

        # Verify they're not live initially
        unpublished_staff.refresh_from_db()
        unpublished_public.refresh_from_db()
        self.assertFalse(unpublished_staff.live)
        self.assertFalse(unpublished_public.live)

        # Create a new active user with matching cnetid
        User.objects.create_user(
            username='newactive', password='password', is_active=True
        )

        # The unpublished pages should now be published automatically
        unpublished_staff.refresh_from_db()
        unpublished_public.refresh_from_db()
        self.assertTrue(unpublished_staff.live)
        self.assertTrue(unpublished_public.live)

    def test_only_live_pages_are_unpublished(self):
        """Test that only live pages are affected when user is deactivated."""
        # Create an already unpublished staff page
        unpublished_staff = StaffPage(
            cnetid='teststaff2',
            slug='test-staff-2',
            title='Test Staff 2',
        )
        self.welcome.add_child(instance=unpublished_staff)
        # Don't publish it

        # Create user and deactivate
        user2 = User.objects.create_user(
            username='teststaff2', password='password', is_active=True
        )
        user2.is_active = False
        user2.save()

        # Page should still be unpublished (no error should occur)
        unpublished_staff.refresh_from_db()
        self.assertFalse(unpublished_staff.live)

    def test_active_user_does_not_unpublish_pages(self):
        """Test that saving an active user doesn't unpublish their pages."""
        # Verify pages are published
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertTrue(self.staff_page.live)
        self.assertTrue(self.public_page.live)

        # Save user without changing active state
        self.user.save()

        # Verify pages are still published
        self.staff_page.refresh_from_db()
        self.public_page.refresh_from_db()
        self.assertTrue(self.staff_page.live)
        self.assertTrue(self.public_page.live)
